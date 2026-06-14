import os
import openpyxl


def parse_excel(file_path: str, sheet_name: str | None = None) -> dict:
    """Read an Excel file and return metadata + rows. Optionally specify sheet name."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {"columns": [], "row_count": 0, "rows": []}

    columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([str(v) if v is not None else "" for v in row])
    wb.close()
    return {"columns": columns, "row_count": len(data_rows), "rows": data_rows}


def write_labeled_excel(columns: list[str], rows: list[list[str]], output_path: str):
    """Write complete data (original columns + AI标签 + AI思考过程) to a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写表头
    for j, col_name in enumerate(columns):
        ws.cell(row=1, column=j + 1, value=col_name)

    # 写数据
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=str(val) if val is not None else "")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()


def write_binarized_excel(columns: list[str], rows: list[list[str]], output_path: str):
    """Write binarized data (original + binary 0/1 columns) to a new Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active

    for j, col_name in enumerate(columns):
        ws.cell(row=1, column=j + 1, value=col_name)

    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
